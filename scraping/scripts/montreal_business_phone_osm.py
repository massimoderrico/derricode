import json
import re
from datetime import date
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUT = Path('/home/massimo/derricode/scraping/data/montreal_business_phone_leads.xlsx')
STATE = OUT.with_name('montreal_business_phone_checkpoint.json')
TODAY = str(date.today())
OVERPASS = 'https://overpass-api.de/api/interpreter'
BBOX = '45.40,-73.75,45.70,-73.45'
TARGET = 1000
HEADERS = [
    'Business Name', 'Category', 'Public Business Phone', 'Address', 'City',
    'Province', 'Postal Code', 'Website URL', 'Source URL', 'OSM Element ID',
    'Size Verification', 'Research Date', 'Notes'
]
EXCLUDED = {
    'place_of_worship', 'school', 'university', 'college', 'social_facility',
    'hospital', 'clinic', 'police', 'fire_station', 'library', 'cemetery',
    'community_centre', 'townhall', 'government', 'public_building',
    'kindergarten', 'childcare', 'grave_yard',
}

def normalized_phone(value):
    return re.sub(r'\D', '', value or '')

def query():
    return f'''[out:json][timeout:60];
    (node["phone"]({BBOX});
    node["contact:phone"]({BBOX}););
    out center tags;'''

def category(tags):
    for key in ('shop', 'office', 'craft', 'amenity', 'tourism', 'leisure', 'industrial'):
        if tags.get(key):
            return f'{key}: {tags[key].replace("_", " ")}'
    return 'business'

def address(tags):
    parts = []
    street = ' '.join(x for x in [tags.get('addr:housenumber'), tags.get('addr:street')] if x)
    if street:
        parts.append(street)
    for key in ('addr:unit', 'addr:city', 'addr:suburb'):
        if tags.get(key):
            parts.append(tags[key])
    return ', '.join(parts) or 'Unknown'

def element_point(element):
    center = element.get('center') or {}
    return element.get('lat', center.get('lat')), element.get('lon', center.get('lon'))

def build_rows(elements):
    rows = []
    seen = set()
    for element in elements:
        tags = element.get('tags') or {}
        name = (tags.get('name') or tags.get('operator') or '').strip()
        phone = (tags.get('phone') or tags.get('contact:phone') or '').strip()
        if not name or not phone or not normalized_phone(phone):
            continue
        if any(tags.get(k) in EXCLUDED for k in ('amenity', 'office', 'tourism', 'leisure')):
            continue
        if tags.get('name') in {'Unknown', 'Unnamed'}:
            continue
        key = (re.sub(r'\W+', '', name).lower(), normalized_phone(phone))
        if key in seen:
            continue
        seen.add(key)
        eid = f"{element['type']}/{element['id']}"
        lat, lon = element_point(element)
        source = f'https://www.openstreetmap.org/{eid}'
        rows.append([
            name,
            category(tags),
            phone,
            address(tags),
            tags.get('addr:city') or tags.get('addr:suburb') or 'Montréal area',
            tags.get('addr:province') or 'Québec',
            tags.get('addr:postcode') or 'Unknown',
            tags.get('website') or tags.get('contact:website') or 'Unknown',
            source,
            eid,
            'Not publicly verified; OSM does not establish employee count or firm size',
            TODAY,
            f'Public OSM business listing with explicit phone tag; coordinates={lat},{lon}.',
        ])
    rows.sort(key=lambda row: (row[0].casefold(), row[2]))
    return rows[:TARGET]

def save(rows, element_count):
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leads'
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:M{ws.max_row}'
    widths = [32, 24, 24, 42, 20, 14, 14, 38, 48, 18, 44, 16, 70]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[1].height = 30
    readme = wb.create_sheet('README')
    notes = [
        ('Purpose', 'Public business phone list for Montreal-area firms, built from OpenStreetMap listings.'),
        ('Exact data rows', str(len(rows))),
        ('Requested target', str(TARGET)),
        ('Source', 'OpenStreetMap Overpass public API; each row retains an OSM source URL.'),
        ('Geography', f'Montreal-area bounding box {BBOX}; node listings may include nearby municipalities.'),
        ('Phone rule', 'A row was included only when a named business listing explicitly exposed a phone or contact:phone tag.'),
        ('Firm size limitation', 'Small/medium status is not verified. No employee count, revenue, or size claim was inferred.'),
        ('Deduplication', 'Deduplicated by normalized business name plus normalized phone.'),
        ('Exclusions', 'Institutional/non-business categories such as schools, hospitals, police, government, and places of worship were excluded where explicitly tagged.'),
        ('Compliance', 'Public business-level data only. No private contacts, owner identities, logins, CAPTCHA bypasses, or guessed numbers.'),
        ('Research date', TODAY),
        ('Overpass element count', str(element_count)),
    ]
    readme.append(['Field', 'Value'])
    for item in notes:
        readme.append(list(item))
    readme.column_dimensions['A'].width = 28
    readme.column_dimensions['B'].width = 130
    for cell in readme[1]:
        cell.fill = fill
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
    for row in readme.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.row > 1:
                cell.font = Font(name='Arial', size=10)
    wb.save(OUT)
    state = {
        'status': 'checkpointed',
        'xlsx': str(OUT),
        'research_date': TODAY,
        'target_rows': TARGET,
        'data_rows': len(rows),
        'overpass_elements': element_count,
        'source': OVERPASS,
    }
    STATE.write_text(json.dumps(state, indent=2))
    return state

def main():
    response = requests.get(
        OVERPASS, params={'data': query()}, timeout=180,
        headers={'User-Agent': 'Derricode public business research/1.0'},
    )
    response.raise_for_status()
    elements = response.json().get('elements', [])
    rows = build_rows(elements)
    if len(rows) < TARGET:
        raise RuntimeError(f'Only {len(rows)} qualifying rows available; refusing to fabricate the requested {TARGET}.')
    state = save(rows, len(elements))
    check = load_workbook(OUT, read_only=True, data_only=True)
    assert check.sheetnames == ['Leads', 'README']
    assert check['Leads'].max_row - 1 == TARGET
    assert all(row[0] and row[2] and row[8] and row[11] for row in check['Leads'].iter_rows(min_row=2, values_only=True))
    print(json.dumps(state, indent=2))

if __name__ == '__main__':
    main()
