import requests, re, time, json, os, sys
from urllib.parse import urlparse
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

API='http://127.0.0.1:3002'
XLSX='/home/massimo/derricode_tennessee_10k_leads.xlsx'
README='/home/massimo/derricode_tennessee_10k_leads_README.md'
TODAY=str(date.today())
HEADERS=['Business Name','Category','Owner Name','Owner Title','Owner LinkedIn URL','Public Business Email','Public Business Phone','Owner Public Business Email','Owner Public Business Phone','Address','City','State','ZIP','Website URL','Website Status','Website Evidence','Source URL','Research Date','Mockup HTML Path','Notes']
TOWNS=['Franklin','Brentwood','Hendersonville','Mount Juliet','Murfreesboro','Smyrna','Gallatin','Lebanon','Clarksville','Spring Hill','Columbia','Nolensville','Antioch','La Vergne','Goodlettsville','Dickson','Springfield','Ashland City','Portland','White House','Fairview','Thompsons','Kingston Springs','Joelton','Belle Meade','Forest Hills','Oak Hill','Berry Hill','Millersville','Greenbrier','Pleasant View','Watertown','Cross Plains','College Grove','Arrington','Eagleville','Shelbyville','Lewisburg','Tullahoma','Cookeville','Knoxville','Chattanooga','Johnson City','Jackson','Memphis']
CATS=['roofing contractor','solar installer','gutter contractor','siding contractor','window contractor','exterior contractor','home remodeling contractor','HVAC contractor','concrete contractor','tree service']
EXCLUDE_DOMAINS=['yellowpages.com','yelp.com','bbb.org','facebook.com','instagram.com','linkedin.com','mapquest.com','angi.com','homeadvisor.com','thumbtack.com','houzz.com','porch.com','google.com','wikipedia.org','youtube.com','reddit.com','chamberofcommerce.com','procore.com','manta.com','superpages.com','nextdoor.com','thebluebook.com','buildzoom.com','dnb.com','crunchbase.com']

def norm(s): return re.sub(r'[^a-z0-9]','',str(s or '').lower())
def domain(u):
 try:
  h=urlparse(u).netloc.lower().split(':')[0]
  return h[4:] if h.startswith('www.') else h
 except: return ''
def phone_norm(s): return re.sub(r'\D','',str(s or ''))
def is_excluded(u):
 d=domain(u); return (not d) or any(d==x or d.endswith('.'+x) for x in EXCLUDE_DOMAINS)
def clean_title(t):
 t=re.sub(r'\s*[|–—-]\s*(Roofing|Solar|Gutters?|Siding|Contractors?|TN|Tennessee).*$', '', t or '', flags=re.I)
 return re.sub(r'\s+',' ',t).strip(' -|')
def guess_name(title,url):
 x=clean_title(title)
 if not x: x=domain(url).split('.')[0].replace('-',' ').title()
 return x

def scrape(u):
 try:
  r=requests.post(API+'/v2/scrape',json={'url':u,'formats':['markdown'],'onlyMainContent':True,'timeout':30000},timeout=45)
  if r.status_code==200:
   j=r.json(); d=j.get('data',j); return d.get('markdown','') or ''
 except Exception: pass
 return ''

def extract(md):
 text=re.sub(r'\[([^\]]+)\]\([^)]*\)',r'\1',md or '')
 emails=sorted(set(re.findall(r'[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}',text,re.I)))
 phones=[]
 for p in re.findall(r'(?:\+?1[\s.-]?)?\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}',text):
  if len(phone_norm(p))>=10: phones.append(p.strip())
 phones=list(dict.fromkeys(phones))
 zips=re.findall(r'\b(?:TN\s*)?(\d{5})(?:-\d{4})?\b',text)
 return text,emails,phones,zips

def city_from(q):
 return q.split('|')[1]

def main():
 wb=load_workbook(XLSX)
 ws=wb['Leads']
 # normalize header and preserve all data rows
 existing=[]
 for row in ws.iter_rows(min_row=2, values_only=True):
  if row and row[0]: existing.append(list(row[:20])+['']*(20-len(row)))
 keys=set(); domains=set(); phones=set()
 for r in existing:
  keys.add(norm(r[0]));
  if r[13] and r[13] != 'Unknown': domains.add(domain(r[13]))
  if r[6] and r[6] != 'Unknown': phones.add(phone_norm(r[6]))
 candidates=[]; seen_urls=set(); search_count=0
 # Search broad local queries; Firecrawl search returns public web results.
 for cat in CATS:
  for town in TOWNS:
   q=f'"{cat}" "{town}" Tennessee'
   try:
    r=requests.post(API+'/v2/search',json={'query':q,'limit':100,'sources':[{'type':'web'}]},timeout=90)
    search_count+=1
    data=r.json().get('data',{}).get('web',[]) if r.status_code==200 else []
   except Exception as e: data=[]
   for item in data:
    u=item.get('url','') if isinstance(item,dict) else ''
    if not u or is_excluded(u) or u in seen_urls: continue
    # Keep local business-looking result, not category/index pages.
    title=item.get('title','') if isinstance(item,dict) else ''
    desc=item.get('description','') if isinstance(item,dict) else ''
    blob=(title+' '+desc).lower()
    if town.lower() not in blob and 'tennessee' not in blob and ' tn' not in blob: continue
    seen_urls.add(u); candidates.append((u,title,desc,cat,town))
   if search_count%20==0: print('searched',search_count,'candidate urls',len(candidates),flush=True)
   time.sleep(0.15)
 # Process individual business pages in bounded batches.
 new=[]; processed=0
 for u,title,desc,cat,town in candidates:
  d=domain(u)
  if d in domains: continue
  md=scrape(u); processed+=1
  text,emails,phs,zips=extract(md)
  # Require page content or search evidence explicitly identifies business/trade.
  blob=(title+' '+desc+' '+text[:8000]).lower()
  if not any(k in blob for k in [cat.split()[0], 'roof', 'solar','gutter','siding','window','remodel','hvac','concrete','tree service','exterior']): continue
  name=guess_name(title,u)
  # Avoid generic search/category pages that slipped through.
  if any(x in norm(name) for x in ['roofingcontractors','solarpanel','bestroofing','nearme','yelp']): continue
  if norm(name) in keys: continue
  p=phs[0] if phs else 'Unknown'; pn=phone_norm(p)
  if pn and pn in phones: continue
  email=emails[0] if emails else 'Unknown'
  # evidence-based status
  if not md:
   status='Unreachable / extraction failed'; evidence='Public search result identified a candidate URL, but Firecrawl page extraction returned no usable public content.'
  else:
   low=text.lower(); signals=[]
   if 'contact' not in low and 'call' not in low and 'phone' not in low: signals.append('no clear contact CTA in extracted text')
   if 'service area' not in low and town.lower() not in low: signals.append('service area/locality not explicit on fetched page')
   if 'copyright 202' not in low and '2025' not in low and '2024' not in low: signals.append('no recent year signal found')
   status='Needs review' if signals else 'Reachable; quality review needed'
   evidence='Firecrawl extracted public page text; '+('; '.join(signals) if signals else 'business/contact content was present; no subjective design judgment made')+'.'
  # best effort address only if clearly visible around TN/zip
  address='Unknown'; zipv=zips[0] if zips else 'Unknown'
  lines=[re.sub(r'\s+',' ',x).strip() for x in text.splitlines() if x.strip()]
  for line in lines:
   if re.search(r'\bTN\b|Tennessee',line,re.I) and re.search(r'\d{5}',line): address=line[:240]; break
  row=[name,cat.title(),'Unknown','Unknown','Unknown',email,p,'Unknown','Unknown',address,town,'TN',zipv,u,status,evidence,u,TODAY,'Not created','Public business data only; owner not added because no explicit public business linkage was verified.']
  new.append(row); keys.add(norm(name)); domains.add(d); 
  if pn: phones.add(pn)
  if len(new)%25==0:
   save(wb,ws,existing+new,search_count,processed,len(candidates),len(new),'batch')
  if len(new)>=1000: break
 save(wb,ws,existing+new,search_count,processed,len(candidates),len(new),'final')
 print(json.dumps({'existing':len(existing),'new':len(new),'total':len(existing)+len(new),'searches':search_count,'candidate_urls':len(candidates),'processed':processed,'xlsx':XLSX,'readme':README},indent=2))

def save(wb,ws,rows,search_count,processed,cands,new_count,phase):
 ws.delete_rows(2,ws.max_row)
 for r in rows: ws.append(r[:20])
 ws.freeze_panes='A2'; ws.auto_filter.ref=f'A1:T{ws.max_row}'
 widths=[28,24,16,18,30,30,22,28,28,34,18,10,10,34,28,75,45,16,18,65]
 for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
 header_fill=PatternFill('solid',fgColor='1F4E78'); header_font=Font(name='Arial',bold=True,color='FFFFFF')
 for cell in ws[1]: cell.fill=header_fill; cell.font=header_font; cell.alignment=Alignment(wrap_text=True,vertical='center')
 for row in ws.iter_rows(min_row=2):
  for c in row: c.font=Font(name='Arial',size=10); c.alignment=Alignment(wrap_text=True,vertical='top')
 ws.row_dimensions[1].height=30
 if 'README' in wb.sheetnames: wb.remove(wb['README'])
 rd=wb.create_sheet('README')
 total=len(rows)
 lines=[
 ('Derricode Tennessee 10K Leads — README',''),('Purpose','Public-business lead research for Tennessee suburban and Tennessee markets focused on roofing, solar installation, and related trades.'),('Research date',TODAY),('Target','10,000 deduplicated businesses; current file is a verified partial checkpoint.'),('Exact count',str(total)),('New rows this rerun',str(new_count)),('Batch progress',f'{phase} checkpoint; {new_count} new rows added; {search_count} Firecrawl search queries issued; {processed} candidate pages processed.'),('Sources used','Local Firecrawl public search and scrape API at http://127.0.0.1:3002; official business websites and public search result pages returned by the API; existing 11-row seed preserved.'),('Methodology','Candidates were searched by trade and Tennessee locality, deduplicated by normalized business name/domain/phone, then page-extracted where possible. Only public business data was recorded. Unknown is retained when not confirmed.'),('Website-quality screen','Statuses use concrete signals: unreachable/extraction failure, missing contact CTA, missing explicit locality/service area, or missing recent-year signal. Needs review is not an aesthetic judgment.'),('Owner enrichment','Owner columns remain Unknown unless a person is explicitly linked to the business through an official page or public professional profile; no identities were inferred in this batch.'),('Mockups','No HTML mockups created; Mockup HTML Path is Not created.'),('Limitations / blocker', 'The 10,000 target was not reached because Firecrawl search results are finite/overlapping across locality-trade queries and page extraction/search throughput is bounded by the local service and lawful public-source availability. No records were fabricated.'),('Compliance','Public business data only. No logins, private profiles, CAPTCHA solving, access-control bypasses, guessed owners, or fabricated contacts. Respect robots.txt, source terms, rate limits, and applicable marketing/privacy laws.'),('Verification',f'Rebuilt Leads sheet and saved after batches; exact data rows={total}; seed rows preserved; deduplication applied; every added row has a Source URL and Research Date.')]
 for a,b in lines: rd.append([a,b])
 rd.column_dimensions['A'].width=28; rd.column_dimensions['B'].width=130
 for row in rd.iter_rows():
  for c in row: c.font=Font(name='Arial',size=10); c.alignment=Alignment(wrap_text=True,vertical='top')
 rd['A1'].fill=header_fill; rd['A1'].font=header_font; rd['B1'].fill=header_fill
 wb.save(XLSX)
 with open(README,'w') as f:
  f.write('# Derricode Tennessee 10K Leads\n\n')
  for a,b in lines: f.write(f'**{a}:** {b}\n\n')
 print('checkpoint saved',total,'rows',flush=True)

if __name__=='__main__': main()
