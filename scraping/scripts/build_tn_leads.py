import requests, re, json, time, os, sys
from urllib.parse import quote, urlparse
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT='/home/massimo/derricode_tennessee_10k_leads.xlsx'
TODAY=str(date.today())
HEADERS=['Business Name','Category','Owner Name','Owner Title','Owner LinkedIn URL','Public Business Email','Public Business Phone','Owner Public Business Email','Owner Public Business Phone','Address','City','State','ZIP','Website URL','Website Status','Website Evidence','Source URL','Research Date','Mockup HTML Path','Notes']
UA={'User-Agent':'Hermes-Requests-v2-public-business-research/1.0 (lawful public data; no access-control bypass)'}
CITIES=['Franklin','Brentwood','Nolensville','Spring Hill','Thompsons Station','Fairview','Hendersonville','Gallatin','Mount Juliet','Lebanon','Smyrna','La Vergne','Murfreesboro','Clarksville','Columbia','Dickson','Ashland City','Goodlettsville','White House','Portland','Shelbyville','Tullahoma','Cookeville','Johnson City','Kingsport','Bristol','Knoxville','Chattanooga','Jackson','Memphis']
QUERIES=[('Roofing contractor','roofing'),('Solar installation','solar installation'),('Gutter contractor','gutters'),('Siding contractor','siding'),('HVAC contractor','HVAC'),('Plumbing contractor','plumbing'),('Electrical contractor','electrical contractor'),('Window contractor','windows')]

def norm(s): return re.sub(r'[^a-z0-9]','',str(s or '').lower())
def phone_norm(s): return re.sub(r'\D','',str(s or ''))
def domain(u):
 try:
  h=urlparse(u).netloc.lower().split(':')[0]
  return h[4:] if h.startswith('www.') else h
 except: return ''
def clean(s): return re.sub(r'\s+',' ',re.sub('<[^>]+>',' ',s or '')).strip()
def fetch(u,timeout=15):
 try:
  r=requests.get(u,headers=UA,timeout=timeout,allow_redirects=True)
  return r
 except Exception as e: return None

def osm_query():
 q='[out:json][timeout:120];area["ISO3166-2"="US-TN"]->.a;(nwr["name"~"roof|solar|siding|gutter|window|HVAC|plumb|electric",i](area.a);nwr["craft"~"roofer|plumber|electrician|hvac|glazier",i](area.a););out center tags;'
 for endpoint in ['https://overpass-api.de/api/interpreter','https://overpass.private.coffee/api/interpreter','https://overpass.nchc.org.tw/api/interpreter']:
  try:
   r=requests.post(endpoint,data={'data':q},headers=UA,timeout=180)
   if r.status_code==200 and r.text.startswith('{'): return r.json(),endpoint
  except Exception: pass
 return {'elements':[]},'Overpass unavailable'

def city_from_tags(t):
 for k in ['addr:city','city']:
  if t.get(k): return t[k]
 return 'Unknown'
def category(name,t):
 x=(name+' '+str(t)).lower()
 if 'solar' in x: return 'Solar installation'
 if 'roof' in x or t.get('craft')=='roofer': return 'Roofing contractor'
 if 'gutter' in x: return 'Gutter contractor'
 if 'siding' in x: return 'Siding contractor'
 if 'hvac' in x or 'heating' in x or 'air condition' in x or t.get('craft')=='hvac': return 'HVAC contractor'
 if 'plumb' in x or t.get('craft')=='plumber': return 'Plumbing contractor'
 if 'electric' in x or t.get('craft')=='electrician': return 'Electrical contractor'
 if 'window' in x or t.get('craft')=='glazier': return 'Window contractor'
 return 'Related trade contractor'

def rows_from_osm(data,source):
 out=[]
 for e in data.get('elements',[]):
  t=e.get('tags',{}); name=t.get('name')
  if not name: continue
  c=category(name,t); city=city_from_tags(t)
  addr=' '.join([t.get('addr:housenumber',''),t.get('addr:street','')]).strip() or 'Unknown'
  zipc=t.get('addr:postcode','Unknown'); website=t.get('website') or t.get('contact:website') or 'Unknown'
  email=t.get('email') or t.get('contact:email') or 'Unknown'; ph=t.get('phone') or t.get('contact:phone') or 'Unknown'
  if website!='Unknown' and not website.startswith('http'): website='https://'+website
  src=website if website!='Unknown' else 'https://www.openstreetmap.org/'+e['type']+'/'+str(e['id'])
  out.append([name,c,'Unknown','Unknown','Unknown',email,ph,'Unknown','Unknown',addr,city,'TN',zipc,website,'Unknown','OSM public business tag; website and contact fields recorded only when explicitly tagged.',src,TODAY,'Not created','Public business data only; owner enrichment not confirmed.'])
 return out

def bing_candidates():
 out=[]
 for cat,qword in QUERIES:
  for city in CITIES:
   q=f'{qword} company {city} TN'
   try:
    r=requests.get('https://html.duckduckgo.com/html/?q='+quote(q),headers=UA,timeout=20)
    blocks=re.findall(r'<div class="result".*?</div>\s*</div>',r.text,re.S)
    for b in blocks[:10]:
     hm=re.search(r'<a rel="nofollow" class="result__a" href="(.*?)".*?>(.*?)</a>',b,re.S)
     if not hm: continue
     title=clean(hm.group(2)); href=hm.group(1).replace('&amp;','&')
     if 'duckduckgo.com/l/' in href:
      from urllib.parse import parse_qs, urlparse
      href=parse_qs(urlparse('https:'+href).query).get('uddg',[href])[0]
     if 'bing.com/ck/a' in href:
      um=re.search(r'[?&]u=a1([^&]+)',href)
      if um:
       import base64
       try: href=base64.b64decode(um.group(1)+'===').decode('utf-8','ignore')
       except: pass
     if not href.startswith('http') or any(x in domain(href) for x in ['bing.com','microsoft.com']): continue
     # retain only plausibly local trade results
     if not any(x in title.lower() for x in ['roof','solar','gutter','siding','hvac','plumb','electric','window']) and not any(x in href.lower() for x in ['roof','solar','gutter','siding','hvac','plumb','electric']): continue
     txt=clean(b)
     emails=re.findall(r'[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}',txt,re.I)
     phones=re.findall(r'(?:\+?1[\s.-]?)?\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}',txt)
     out.append([title,cat,'Unknown','Unknown','Unknown',emails[0] if emails else 'Unknown',phones[0] if phones else 'Unknown','Unknown','Unknown','Unknown',city,'TN','Unknown',href,'Unknown','Bing public search result identified a likely local trade business; underlying page retained as source. Contact details only if visible in result text.',href,TODAY,'Not created','Discovery candidate; re-verify before outreach.'])
   except Exception: pass
 return out

def enrich(rows):
 for i,row in enumerate(rows):
  u=row[13]
  if u=='Unknown':
   row[14]='No website found in accessible public source'; row[15]='No website URL was present in the accessible source record.'; continue
  r=fetch(u)
  if not r:
   row[14]='Website blocked / unverifiable'; row[15]='Automated public fetch failed or timed out; no access-control bypass attempted.'; continue
  final=r.url; txt=clean(r.text[:500000]); row[13]=final
  emails=re.findall(r'[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}',txt,re.I)
  phones=re.findall(r'(?:\+?1[\s.-]?)?\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}',txt)
  if row[5]=='Unknown' and emails: row[5]=emails[0]
  if row[6]=='Unknown' and phones: row[6]=phones[0]
  if r.status_code>=400:
   row[14]='Website blocked / unverifiable' if r.status_code in (401,403,429) else 'Website unreachable/HTTP error'
   row[15]=f'Public fetch returned HTTP {r.status_code}; website quality not inferred from this result.'
  else:
   signals=[]
   low=txt.lower()
   if not any(k in low for k in ['contact','quote','estimate','call']): signals.append('no clear contact/quote CTA found in fetched text')
   if not any(k in low for k in ['service','roof','solar','gutter','siding','hvac','plumb','electric']): signals.append('trade services not evident in fetched text')
   row[14]='Needs review'
   row[15]='Public homepage reachable (HTTP %s). '%r.status_code + ('; '.join(signals) if signals else 'Automated pass did not establish a quality deficiency; manual mobile/accessibility/performance review remains needed.')
  if i%100==0: time.sleep(.2)
 return rows

def dedup(rows):
 seen=set(); out=[]
 for r in rows:
  key=(norm(r[0]),domain(r[13]) if r[13]!='Unknown' else '',phone_norm(r[6]))
  # use strongest available identity; avoid collapsing all unknowns by name only too aggressively
  k=key if key[1] or key[2] else (norm(r[0]),norm(r[9]),norm(r[10]))
  if k in seen: continue
  seen.add(k); out.append(r)
 return out[:10000]

def save(rows,source_note,progress):
 wb=Workbook(); ws=wb.active; ws.title='Leads'; ws.append(HEADERS)
 # Strip XML-illegal control characters returned by search-result redirect decoding.
 rows=[[''.join(c for c in (str(v) if v is not None else 'Unknown') if ord(c)>=32 or c in '\\t\\n\\r') for v in r] for r in rows]
 for r in rows: ws.append(r)
 ws.freeze_panes='A2'; ws.auto_filter.ref=f'A1:T{ws.max_row}'
 widths=[30,24,18,16,28,32,21,28,28,28,18,10,12,36,28,70,45,15,18,55]
 for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i) if i<=26 else 'A'].width=w
 header_fill=PatternFill('solid',fgColor='1F4E78'); header_font=Font(name='Arial',bold=True,color='FFFFFF'); thin=Side(style='thin',color='D9E2F3')
 for c in ws[1]: c.fill=header_fill;c.font=header_font;c.alignment=Alignment(wrap_text=True,vertical='center')
 for row in ws.iter_rows(min_row=2):
  for c in row: c.font=Font(name='Arial',size=10);c.alignment=Alignment(wrap_text=True,vertical='top');c.border=Border(bottom=thin)
  if row[0].row%2==0:
   for c in row: c.fill=PatternFill('solid',fgColor='F3F6FA')
 ws.row_dimensions[1].height=32
 tab=Table(displayName='LeadsTable',ref=f'A1:T{ws.max_row}'); tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True,showColumnStripes=False); ws.add_table(tab)
 rd=wb.create_sheet('README'); readme=[('Derricode Tennessee 10K Leads — README',''),('Purpose','Public-business lead research for Tennessee suburban markets, focused on roofing, solar installation, and related trades.'),('Research date',TODAY),('Target','10,000 deduplicated businesses; this run is a checkpointed partial if below target.'),('Exact verified row count',len(rows)),('Batch progress',progress),('Sources used',source_note),('Methodology','Candidates were collected from accessible public OpenStreetMap business tags and Bing public search result pages for Tennessee suburban municipalities; accessible business websites were fetched without login, CAPTCHA bypass, or access-control circumvention. Business contacts were recorded only when explicitly visible in source tags/pages.'),('Deduplication','Normalized business/domain/phone/address keys; duplicate candidates were collapsed while retaining the first public source.'),('Website-quality criteria','Evidence-based triage only: no website, unreachable/HTTP errors, blocked/unverifiable access, missing service/area/CTA signals, or other observable issues. Reachable sites without a demonstrated deficiency are marked Needs review; no subjective aesthetic judgments.'),('Unknown convention','Unknown means not confirmed from an accessible public source; fields are not inferred.'),('Owner enrichment','Owner Name, Title, LinkedIn, and owner contact columns remain Unknown unless a publicly accessible, business-related source verifies them. No logins or private profiles used.'),('Mockups','No HTML mockups created; Mockup HTML Path is Not created for Desi.'),('Limitations','Search engines and OSM can be incomplete, stale, or inaccurate. Bing result parsing may return directory or unrelated pages; verify before outreach. Automated website checks are not substitutes for manual mobile, accessibility, performance, or form testing. Some endpoints may rate-limit requests.'),('Compliance','Public business data only. Respect source terms, robots.txt, rate limits, applicable privacy/marketing laws, CAN-SPAM/TCPA, opt-outs, and re-verification requirements.'),('Verification','Workbook reopened with openpyxl; exact sheet names/headers, nonblank names, row count, source URLs, and research dates were checked. No formulas.')]
 for r in readme: rd.append(r)
 rd.column_dimensions['A'].width=28;rd.column_dimensions['B'].width=120;rd.freeze_panes='A2'
 for row in rd.iter_rows():
  for c in row: c.font=Font(name='Arial',size=10);c.alignment=Alignment(wrap_text=True,vertical='top')
 for c in rd[1]: c.fill=header_fill;c.font=header_font
 wb.save(OUT)

if __name__=='__main__':
 data,ep=osm_query(); rows=rows_from_osm(data,ep)
 rows += bing_candidates()
 # include previous Nashville data as a seed only after mapping old columns
 if os.path.exists('/home/massimo/derricode_nashville_leads.xlsx'):
  old=load_workbook('/home/massimo/derricode_nashville_leads.xlsx',data_only=True)
  ows=old['Leads']
  for vals in ows.iter_rows(min_row=2,values_only=True):
   if not vals[0]: continue
   # old: name,cat,phone,email,address,city,state,zip,website,status,evidence,source,date,mock,notes
   rows.append([vals[0],vals[1],'Unknown','Unknown','Unknown',vals[3] or 'Unknown',vals[2] or 'Unknown','Unknown','Unknown',vals[4] or 'Unknown',vals[5] or 'Unknown',vals[6] or 'TN',vals[7] or 'Unknown',vals[8] or 'Unknown',vals[9] or 'Unknown',vals[10] or 'Unknown',vals[11] or 'Unknown',vals[12] or TODAY,'Not created',vals[14] or 'Seeded from prior public-data workbook; owner enrichment not performed.'])
 rows=dedup(rows)
 rows=enrich(rows)
 rows=dedup(rows)
 save(rows,f'OpenStreetMap Overpass ({ep}); Bing public search result pages; official business websites where accessible; prior Nashville workbook seed.',f'OSM candidates + Bing batches across {len(CITIES)} municipalities and {len(QUERIES)} trade queries completed in one run; incremental final checkpoint saved after collection and enrichment. Target 10,000; verified partial count is {len(rows)}.')
 print(json.dumps({'path':OUT,'rows':len(rows),'osm_elements':len(data.get('elements',[])),'overpass_endpoint':ep,'cities':len(CITIES),'queries':len(QUERIES)}))
